"""
ui/spreadsheet_widget.py — Visualizzatore/editor foglio di calcolo per NotePadPQ.
Supporta: CSV (con wizard import), XLSX, XLS (sola lettura), ODS.
"""
from __future__ import annotations

import csv
import re as _re
from pathlib import Path
from typing import Any, Optional

from PyQt6.QtCore import (
    Qt, QAbstractTableModel, QModelIndex, QSortFilterProxyModel, QTimer,
    pyqtSignal,
)
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import (
    QAbstractItemView, QApplication, QButtonGroup, QCheckBox, QComboBox,
    QDialog, QDialogButtonBox, QFileDialog, QGroupBox, QHBoxLayout, QHeaderView,
    QInputDialog, QLabel, QLineEdit, QMenu, QMessageBox, QPlainTextEdit,
    QPushButton, QRadioButton, QSizePolicy, QTableView, QTableWidget,
    QTableWidgetItem, QVBoxLayout, QWidget,
)


# ─── Formula helpers ─────────────────────────────────────────────────────────

def _to_num(v) -> float:
    if isinstance(v, bool):  return 1.0 if v else 0.0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        try: return float(v.replace(",", "."))
        except ValueError: return 0.0
    return 0.0

def _is_numlike(v) -> bool:
    if isinstance(v, bool): return True
    if isinstance(v, (int, float)): return True
    if isinstance(v, str):
        try: float(v.replace(",", ".")); return True
        except ValueError: return False
    return False

def _to_bool(v) -> bool:
    if isinstance(v, bool): return v
    if isinstance(v, (int, float)): return v != 0
    if isinstance(v, str):
        if v.upper() in ("TRUE", "VERO"):  return True
        if v.upper() in ("FALSE", "FALSO"): return False
        try: return float(v.replace(",", ".")) != 0
        except ValueError: return bool(v)
    return bool(v)

def _safe_avg(vals: list) -> float:
    nums = [_to_num(v) for v in vals if _is_numlike(v)]
    return sum(nums) / len(nums) if nums else 0.0

_FORMULA_FUNCS: dict = {
    "SUM":          lambda v: sum(_to_num(x) for x in v if _is_numlike(x)),
    "AVERAGE":      lambda v: _safe_avg(v),
    "AVG":          lambda v: _safe_avg(v),
    "MIN":          lambda v: min((_to_num(x) for x in v if _is_numlike(x)), default=0),
    "MAX":          lambda v: max((_to_num(x) for x in v if _is_numlike(x)), default=0),
    "COUNT":        lambda v: sum(1 for x in v if _is_numlike(x)),
    "COUNTA":       lambda v: sum(1 for x in v if x != "" and x is not None),
    "IF":           lambda v: v[1] if _to_bool(v[0]) else (v[2] if len(v) > 2 else ""),
    "ABS":          lambda v: abs(_to_num(v[0])) if v else 0,
    "ROUND":        lambda v: round(_to_num(v[0]), int(_to_num(v[1])) if len(v) > 1 else 0),
    "SQRT":         lambda v: _to_num(v[0]) ** 0.5 if v else 0,
    "INT":          lambda v: int(_to_num(v[0])) if v else 0,
    "LEN":          lambda v: len(str(v[0])) if v else 0,
    "CONCAT":       lambda v: "".join(str(x) for x in v),
    "CONCATENATE":  lambda v: "".join(str(x) for x in v),
    "UPPER":        lambda v: str(v[0]).upper() if v else "",
    "LOWER":        lambda v: str(v[0]).lower() if v else "",
    "TRIM":         lambda v: str(v[0]).strip() if v else "",
    "LEFT":         lambda v: str(v[0])[:int(_to_num(v[1]))] if len(v) > 1 else str(v[0])[:1],
    "RIGHT":        lambda v: str(v[0])[-int(_to_num(v[1])):] if len(v) > 1 else str(v[0])[-1:],
    "MID":          lambda v: str(v[0])[int(_to_num(v[1]))-1 : int(_to_num(v[1]))-1+int(_to_num(v[2]))] if len(v) > 2 else "",
}

_CELL_PAT = _re.compile(r'^\$?([A-Za-z]+)\$?(\d+)$')

# Categorie per il menu "Inserisci funzione"
_FORMULA_MENU = [
    ("Matematica", [
        ("SUM(range)",         "=SUM(",        "Somma tutti i valori nel range\nEs: =SUM(A1:A10)"),
        ("AVERAGE(range)",     "=AVERAGE(",    "Media aritmetica dei valori\nEs: =AVERAGE(B1:B5)"),
        ("MIN(range)",         "=MIN(",        "Valore minimo nel range\nEs: =MIN(C1:C100)"),
        ("MAX(range)",         "=MAX(",        "Valore massimo nel range\nEs: =MAX(C1:C100)"),
        ("COUNT(range)",       "=COUNT(",      "Conta le celle con valore numerico\nEs: =COUNT(A1:A50)"),
        ("COUNTA(range)",      "=COUNTA(",     "Conta le celle non vuote\nEs: =COUNTA(A1:A50)"),
        ("ABS(n)",             "=ABS(",        "Valore assoluto\nEs: =ABS(A1)"),
        ("ROUND(n, dec)",      "=ROUND(",      "Arrotonda a n cifre decimali\nEs: =ROUND(A1,2)"),
        ("SQRT(n)",            "=SQRT(",       "Radice quadrata\nEs: =SQRT(A1)"),
        ("INT(n)",             "=INT(",        "Parte intera (tronca verso zero)\nEs: =INT(3.7) → 3"),
    ]),
    ("Testo", [
        ("LEN(testo)",          "=LEN(",        "Numero di caratteri della stringa\nEs: =LEN(A1)"),
        ("CONCAT(a, b, …)",     "=CONCAT(",     "Unisce due o più stringhe\nEs: =CONCAT(A1,\" \",B1)"),
        ("UPPER(testo)",        "=UPPER(",      "Converte in MAIUSCOLO\nEs: =UPPER(A1)"),
        ("LOWER(testo)",        "=LOWER(",      "Converte in minuscolo\nEs: =LOWER(A1)"),
        ("TRIM(testo)",         "=TRIM(",       "Rimuove spazi iniziali e finali\nEs: =TRIM(A1)"),
        ("LEFT(testo, n)",      "=LEFT(",       "Primi n caratteri da sinistra\nEs: =LEFT(A1,3)"),
        ("RIGHT(testo, n)",     "=RIGHT(",      "Ultimi n caratteri da destra\nEs: =RIGHT(A1,4)"),
        ("MID(testo, start, n)","=MID(",        "Sottostringa: n caratteri da posizione start\nEs: =MID(A1,2,5)"),
    ]),
    ("Logica", [
        ("IF(cond, vero, falso)", "=IF(",       "Se cond è vera restituisce vero, altrimenti falso\nEs: =IF(A1>0,\"positivo\",\"negativo\")"),
    ]),
]

def _col_letter(n: int) -> str:
    """Converte indice 0-based in lettere colonna stile Excel: 0→A, 25→Z, 26→AA."""
    result = ""
    n += 1
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


# ─── FormulaEngine ────────────────────────────────────────────────────────────

class FormulaEngine:
    """Valutatore di formule stile Excel per SpreadsheetModel.

    Supporta: riferimenti cella (A1), range (A1:B3), funzioni (SUM, IF, …),
    operatori aritmetici, confronto e concatenazione stringa (&).
    """

    @staticmethod
    def _col_idx(letters: str) -> int:
        r = 0
        for c in letters.upper():
            r = r * 26 + (ord(c) - 64)
        return r - 1

    def __init__(self, model: "SpreadsheetModel"):
        self._model = model

    def evaluate(self, formula: str, visited: frozenset = frozenset()) -> str:
        try:
            tokens = self._tokenize(formula[1:])  # strip leading =
            pos = [0]
            result = self._parse_expr(tokens, pos, visited)
            return self._fmt(result)
        except ZeroDivisionError:
            return "#DIV/0!"
        except RecursionError:
            return "#REF!"
        except Exception:
            return "#ERRORE"

    # ── Formatter ─────────────────────────────────────────────────────────────

    @staticmethod
    def _fmt(val) -> str:
        if isinstance(val, bool):
            return "VERO" if val else "FALSO"
        if isinstance(val, float):
            if val == int(val) and abs(val) < 1e15:
                return str(int(val))
            return f"{val:.10g}"
        return str(val) if val is not None else ""

    # ── Cell / range access ───────────────────────────────────────────────────

    def _cell_val(self, col: int, row: int, visited: frozenset):
        key = (row, col)
        if key in visited:
            raise RecursionError
        m = self._model
        if row >= m.rowCount() or col >= m.columnCount():
            return ""
        raw = m._data[row][col] if col < len(m._data[row]) else ""
        if not raw:
            return ""
        if raw.startswith("="):
            result = self.evaluate(raw, visited | {key})
            if result.startswith("#"):
                return result
            try:
                return float(result.replace(",", "."))
            except ValueError:
                return result
        try:
            return float(raw.replace(",", "."))
        except ValueError:
            return raw

    def _range_vals(self, c1: int, r1: int, c2: int, r2: int, visited: frozenset) -> list:
        return [
            self._cell_val(c, r, visited)
            for r in range(min(r1, r2), max(r1, r2) + 1)
            for c in range(min(c1, c2), max(c1, c2) + 1)
        ]

    # ── Tokenizer ─────────────────────────────────────────────────────────────

    @staticmethod
    def _tokenize(s: str) -> list:
        toks: list = []
        i = 0
        while i < len(s):
            if s[i].isspace():
                i += 1
                continue
            if s[i] == '"':
                j = i + 1
                while j < len(s) and s[j] != '"':
                    j += 1
                toks.append(("STR", s[i+1:j]))
                i = j + 1
                continue
            if s[i].isdigit() or (s[i] == '.' and i+1 < len(s) and s[i+1].isdigit()):
                j = i
                while j < len(s) and (s[j].isdigit() or s[j] == '.'):
                    j += 1
                toks.append(("NUM", float(s[i:j])))
                i = j
                continue
            if s[i].isalpha() or s[i] == '$':
                j = i
                while j < len(s) and (s[j].isalnum() or s[j] in ('$', '_')):
                    j += 1
                word = s[i:j].replace('$', '')
                if _CELL_PAT.match(word):
                    toks.append(("CELL", word.upper()))
                else:
                    toks.append(("IDENT", word.upper()))
                i = j
                continue
            if i + 1 < len(s) and s[i:i+2] in ('<>', '<=', '>='):
                toks.append(("OP", s[i:i+2]))
                i += 2
                continue
            if s[i] in '+-*/^&':
                toks.append(("OP", s[i]))
            elif s[i] in '(),:;':
                toks.append(("PUNCT", s[i]))
            elif s[i] in '=<>':
                toks.append(("OP", s[i]))
            i += 1
        toks.append(("EOF", None))
        return toks

    # ── Recursive-descent parser ──────────────────────────────────────────────

    def _parse_expr(self, toks: list, pos: list, visited: frozenset):
        return self._parse_compare(toks, pos, visited)

    def _parse_compare(self, toks, pos, visited):
        left = self._parse_concat(toks, pos, visited)
        if toks[pos[0]][0] == 'OP' and toks[pos[0]][1] in ('=', '<>', '<', '>', '<=', '>='):
            op = toks[pos[0]][1]; pos[0] += 1
            right = self._parse_concat(toks, pos, visited)
            lc = _to_num(left) if _is_numlike(left) else str(left).lower()
            rc = _to_num(right) if _is_numlike(right) else str(right).lower()
            return {'=': lc==rc, '<>': lc!=rc, '<': lc<rc, '>': lc>rc,
                    '<=': lc<=rc, '>=': lc>=rc}[op]
        return left

    def _parse_concat(self, toks, pos, visited):
        left = self._parse_add(toks, pos, visited)
        while toks[pos[0]] == ('OP', '&'):
            pos[0] += 1
            right = self._parse_add(toks, pos, visited)
            left = str(left) + str(right)
        return left

    def _parse_add(self, toks, pos, visited):
        left = self._parse_mul(toks, pos, visited)
        while toks[pos[0]][0] == 'OP' and toks[pos[0]][1] in ('+', '-'):
            op = toks[pos[0]][1]; pos[0] += 1
            right = self._parse_mul(toks, pos, visited)
            left = (_to_num(left) + _to_num(right)) if op == '+' else (_to_num(left) - _to_num(right))
        return left

    def _parse_mul(self, toks, pos, visited):
        left = self._parse_unary(toks, pos, visited)
        while toks[pos[0]][0] == 'OP' and toks[pos[0]][1] in ('*', '/'):
            op = toks[pos[0]][1]; pos[0] += 1
            right = self._parse_unary(toks, pos, visited)
            if op == '/':
                rv = _to_num(right)
                left = _to_num(left) / rv if rv != 0 else "#DIV/0!"
            else:
                left = _to_num(left) * _to_num(right)
        return left

    def _parse_unary(self, toks, pos, visited):
        if toks[pos[0]] == ('OP', '-'):
            pos[0] += 1
            return -_to_num(self._parse_power(toks, pos, visited))
        if toks[pos[0]] == ('OP', '+'):
            pos[0] += 1
            return _to_num(self._parse_power(toks, pos, visited))
        return self._parse_power(toks, pos, visited)

    def _parse_power(self, toks, pos, visited):
        base = self._parse_primary(toks, pos, visited)
        if toks[pos[0]] == ('OP', '^'):
            pos[0] += 1
            exp = self._parse_primary(toks, pos, visited)
            return _to_num(base) ** _to_num(exp)
        return base

    def _parse_primary(self, toks, pos, visited):
        tok = toks[pos[0]]

        if tok[0] == 'NUM':
            pos[0] += 1; return tok[1]

        if tok[0] == 'STR':
            pos[0] += 1; return tok[1]

        if tok[0] == 'PUNCT' and tok[1] == '(':
            pos[0] += 1
            val = self._parse_expr(toks, pos, visited)
            if toks[pos[0]] == ('PUNCT', ')'):
                pos[0] += 1
            return val

        if tok[0] == 'IDENT':
            name = tok[1]; pos[0] += 1
            if name == 'TRUE':  return True
            if name == 'FALSE': return False
            if name == 'VERO':  return True
            if name == 'FALSO': return False
            # Function call
            if toks[pos[0]] == ('PUNCT', '('):
                pos[0] += 1
                args = self._parse_args(toks, pos, visited)
                if toks[pos[0]] == ('PUNCT', ')'):
                    pos[0] += 1
                fn = _FORMULA_FUNCS.get(name)
                if fn is None:
                    return f"#NOME?"
                flat = [x for a in args for x in (a if isinstance(a, list) else [a])]
                return fn(flat)
            return 0.0  # bare identifier without ()

        if tok[0] == 'CELL':
            cell_str = tok[1]; pos[0] += 1
            # Range?
            if toks[pos[0]] == ('PUNCT', ':'):
                pos[0] += 1
                cell2 = toks[pos[0]][1]; pos[0] += 1
                return self._resolve_range(cell_str, cell2, visited)
            return self._resolve_cell(cell_str, visited)

        pos[0] += 1  # skip unknown
        return 0.0

    def _parse_args(self, toks, pos, visited) -> list:
        args = []
        if toks[pos[0]] == ('PUNCT', ')'):
            return args
        args.append(self._parse_arg(toks, pos, visited))
        while toks[pos[0]][0] == 'PUNCT' and toks[pos[0]][1] in (',', ';'):
            pos[0] += 1
            args.append(self._parse_arg(toks, pos, visited))
        return args

    def _parse_arg(self, toks, pos, visited):
        if toks[pos[0]][0] == 'CELL':
            saved = pos[0]
            cell1 = toks[pos[0]][1]; pos[0] += 1
            if toks[pos[0]] == ('PUNCT', ':'):
                pos[0] += 1
                cell2 = toks[pos[0]][1]; pos[0] += 1
                return self._resolve_range(cell1, cell2, visited)
            pos[0] = saved
        return self._parse_expr(toks, pos, visited)

    # ── Cell / range resolution ───────────────────────────────────────────────

    def _resolve_cell(self, cell_str: str, visited: frozenset):
        m = _CELL_PAT.match(cell_str)
        if not m:
            return 0.0
        col = self._col_idx(m.group(1))
        row = int(m.group(2)) - 1
        return self._cell_val(col, row, visited)

    def _resolve_range(self, c1: str, c2: str, visited: frozenset) -> list:
        m1 = _CELL_PAT.match(c1)
        m2 = _CELL_PAT.match(c2)
        if not m1 or not m2:
            return []
        return self._range_vals(
            self._col_idx(m1.group(1)), int(m1.group(2)) - 1,
            self._col_idx(m2.group(1)), int(m2.group(2)) - 1,
            visited
        )


# ─── ImportWizardDialog ───────────────────────────────────────────────────────

class ImportWizardDialog(QDialog):
    """Wizard stile Excel per scegliere delimitatore, encoding e riga intestazione."""

    def __init__(self, path: Path, parent=None):
        super().__init__(parent)
        self._path = path
        self._raw_lines: list[str] = []
        self._encoding = "utf-8-sig"
        self._parsed_headers: list[str] = []
        self._parsed_data: list[list[str]] = []

        self.setWindowTitle(f"Importa: {path.name}")
        self.setMinimumSize(720, 560)
        self.resize(820, 620)

        self._detect_encoding()
        self._load_raw_lines()
        self._build_ui()
        self._auto_select_delimiter()
        self._update_preview()

    def _detect_encoding(self) -> None:
        try:
            import chardet
            raw = self._path.read_bytes()
            detected = chardet.detect(raw)
            self._encoding = detected.get("encoding") or "utf-8"
        except (ImportError, Exception):
            self._encoding = "utf-8-sig"

    def _load_raw_lines(self) -> None:
        try:
            with open(self._path, "r", encoding=self._encoding, errors="replace") as f:
                self._raw_lines = [f.readline() for _ in range(60)]
            self._raw_lines = [ln.rstrip("\n\r") for ln in self._raw_lines if ln.strip()]
        except Exception:
            self._raw_lines = []

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        grp_raw = QGroupBox("Anteprima testo grezzo (prime 15 righe)")
        raw_lay = QVBoxLayout(grp_raw)
        self._raw_view = QPlainTextEdit()
        self._raw_view.setReadOnly(True)
        self._raw_view.setMaximumHeight(110)
        self._raw_view.setFont(QFont("Monospace", 9))
        self._raw_view.setPlainText("\n".join(self._raw_lines[:15]))
        raw_lay.addWidget(self._raw_view)
        layout.addWidget(grp_raw)

        opts_row = QHBoxLayout()

        grp_delim = QGroupBox("Separatore di colonna")
        delim_lay = QVBoxLayout(grp_delim)
        self._delim_group = QButtonGroup(self)

        for label, char in [
            ("Virgola  ,", ","), ("Punto e virgola  ;", ";"),
            ("Tab  \\t", "\t"), ("Pipe  |", "|"),
            ("Backslash  \\", "\\"), ("Spazio   ", " "),
        ]:
            rb = QRadioButton(label)
            rb.setProperty("delim_char", char)
            rb.toggled.connect(self._update_preview)
            self._delim_group.addButton(rb)
            delim_lay.addWidget(rb)

        rb_custom = QRadioButton("Altro:")
        rb_custom.setProperty("delim_char", "__custom__")
        rb_custom.toggled.connect(self._update_preview)
        self._delim_group.addButton(rb_custom)
        custom_row = QHBoxLayout()
        custom_row.addWidget(rb_custom)
        self._custom_delim = QLineEdit()
        self._custom_delim.setMaximumWidth(40)
        self._custom_delim.setMaxLength(1)
        self._custom_delim.textChanged.connect(self._update_preview)
        custom_row.addWidget(self._custom_delim)
        custom_row.addStretch()
        delim_lay.addLayout(custom_row)
        grp_delim.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Minimum)
        opts_row.addWidget(grp_delim)

        grp_opts = QGroupBox("Opzioni aggiuntive")
        other_lay = QVBoxLayout(grp_opts)
        self._cb_header = QCheckBox("Prima riga come intestazione colonne")
        self._cb_header.setChecked(True)
        self._cb_header.toggled.connect(self._update_preview)
        other_lay.addWidget(self._cb_header)
        enc_row = QHBoxLayout()
        enc_row.addWidget(QLabel("Codifica testo:"))
        self._enc_combo = QComboBox()
        _ENCS = ["utf-8", "utf-8-sig", "latin-1", "windows-1252", "utf-16", "iso-8859-1"]
        self._enc_combo.addItems(_ENCS)
        dl = self._encoding.lower().replace("-", "")
        for i, e in enumerate(_ENCS):
            if e.replace("-", "") == dl:
                self._enc_combo.setCurrentIndex(i)
                break
        self._enc_combo.currentIndexChanged.connect(self._on_encoding_changed)
        enc_row.addWidget(self._enc_combo)
        enc_row.addStretch()
        other_lay.addLayout(enc_row)
        other_lay.addWidget(QLabel("(rilevata automaticamente)"))
        other_lay.addStretch()
        opts_row.addWidget(grp_opts)
        layout.addLayout(opts_row)

        grp_prev = QGroupBox("Anteprima tabella (prime 10 righe dati)")
        prev_lay = QVBoxLayout(grp_prev)
        self._preview_table = QTableWidget()
        self._preview_table.setMaximumHeight(180)
        self._preview_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._preview_table.setAlternatingRowColors(True)
        prev_lay.addWidget(self._preview_table)
        layout.addWidget(grp_prev)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Importa")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _auto_select_delimiter(self) -> None:
        if not self._raw_lines:
            return
        sample = "\n".join(self._raw_lines[:5])
        counts = {",": sample.count(","), ";": sample.count(";"),
                  "\t": sample.count("\t"), "|": sample.count("|"),
                  " ": sample.count(" ")}
        best = max(counts, key=counts.get)
        idx = {",": 0, ";": 1, "\t": 2, "|": 3, "\\": 4, " ": 5}.get(best, 0)
        buttons = self._delim_group.buttons()
        if idx < len(buttons):
            buttons[idx].setChecked(True)

    def _get_delimiter(self) -> str:
        for btn in self._delim_group.buttons():
            if btn.isChecked():
                d = btn.property("delim_char")
                return self._custom_delim.text() or "," if d == "__custom__" else d
        return ","

    def _on_encoding_changed(self) -> None:
        self._encoding = self._enc_combo.currentText()
        self._load_raw_lines()
        self._raw_view.setPlainText("\n".join(self._raw_lines[:15]))
        self._update_preview()

    def _update_preview(self) -> None:
        delim = self._get_delimiter()
        first_row_header = self._cb_header.isChecked()
        rows: list[list[str]] = []
        for line in self._raw_lines[:20]:
            if not line:
                continue
            try:
                parsed = next(csv.reader([line], delimiter=delim))
            except Exception:
                parsed = [line]
            rows.append(parsed)
        if not rows:
            self._parsed_headers = []
            self._parsed_data = []
            self._preview_table.clear()
            return
        if first_row_header:
            self._parsed_headers = rows[0]
            self._parsed_data = rows[1:]
        else:
            n = max((len(r) for r in rows), default=1)
            self._parsed_headers = [f"Col{i+1}" for i in range(n)]
            self._parsed_data = rows
        n_cols = len(self._parsed_headers)
        n_rows = min(len(self._parsed_data), 10)
        self._preview_table.clear()
        self._preview_table.setColumnCount(n_cols)
        self._preview_table.setRowCount(n_rows)
        self._preview_table.setHorizontalHeaderLabels(self._parsed_headers)
        for r, row in enumerate(self._parsed_data[:10]):
            for c in range(n_cols):
                val = row[c] if c < len(row) else ""
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self._preview_table.setItem(r, c, item)
        self._preview_table.resizeColumnsToContents()

    def get_settings(self) -> tuple[str, str, bool]:
        return self._get_delimiter(), self._encoding, self._cb_header.isChecked()


# ─── _FilterProxy ─────────────────────────────────────────────────────────────

class _FilterProxy(QSortFilterProxyModel):
    """Filtro testo su colonna specifica o su tutte le colonne."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._col: int = -1    # -1 = tutte le colonne
        self._text: str = ""
        self.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.setDynamicSortFilter(True)

    def set_filter(self, col: int, text: str) -> None:
        self._col = col
        self._text = text.strip().lower()
        self.invalidateFilter()

    def clear(self) -> None:
        self._col = -1
        self._text = ""
        self.invalidateFilter()

    def active_col(self) -> int:
        return self._col

    def active_text(self) -> str:
        return self._text

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        if not self._text:
            return True
        model = self.sourceModel()
        if self._col == -1:
            for c in range(model.columnCount()):
                idx = model.index(source_row, c, source_parent)
                val = (model.data(idx, Qt.ItemDataRole.DisplayRole) or "").lower()
                if self._text in val:
                    return True
            return False
        idx = model.index(source_row, self._col, source_parent)
        val = (model.data(idx, Qt.ItemDataRole.DisplayRole) or "").lower()
        return self._text in val

    def lessThan(self, left: QModelIndex, right: QModelIndex) -> bool:
        return False


# ─── SpreadsheetModel ─────────────────────────────────────────────────────────

class SpreadsheetModel(QAbstractTableModel):
    modified_changed = pyqtSignal(bool)

    def __init__(self, headers: list[str], data: list[list[str]], parent=None):
        super().__init__(parent)
        self._headers: list[str] = headers[:]
        self._data: list[list[str]] = [row[:] for row in data]
        self._modified = False
        self._sort_keys: list[tuple[int, Qt.SortOrder]] = []
        self._engine: Optional[FormulaEngine] = None

    def _get_engine(self) -> "FormulaEngine":
        if self._engine is None:
            self._engine = FormulaEngine(self)
        return self._engine

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return len(self._data)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        if self._headers:
            return len(self._headers)
        return max((len(r) for r in self._data), default=0)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole) -> Any:
        if not index.isValid():
            return None
        r, c = index.row(), index.column()
        if r >= len(self._data):
            return None
        row = self._data[r]
        raw = row[c] if c < len(row) else ""
        if role == Qt.ItemDataRole.EditRole:
            return raw  # restituisce la formula grezza per la modifica
        is_formula = isinstance(raw, str) and raw.startswith("=")
        if role == Qt.ItemDataRole.DisplayRole:
            if is_formula:
                return self._get_engine().evaluate(raw)
            return raw
        if role == Qt.ItemDataRole.ForegroundRole and is_formula:
            from PyQt6.QtGui import QColor
            return QColor("#5baaff")  # formule in azzurro
        if role == Qt.ItemDataRole.TextAlignmentRole:
            display = self._get_engine().evaluate(raw) if is_formula else raw
            try:
                float(str(display).replace(",", "."))
                return int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            except (ValueError, AttributeError):
                return int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        return None

    def headerData(self, section: int, orientation: Qt.Orientation,
                   role: int = Qt.ItemDataRole.DisplayRole) -> Any:
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == Qt.Orientation.Horizontal:
            name = self._headers[section] if section < len(self._headers) else str(section + 1)
            letter = _col_letter(section)
            label = f"{letter}  {name}"
            # Aggiunge indicatori di ordinamento
            for priority, (col, order) in enumerate(self._sort_keys):
                if col == section:
                    arrow = "↑" if order == Qt.SortOrder.AscendingOrder else "↓"
                    suffix = f" {arrow}" if len(self._sort_keys) == 1 else f" {arrow}{priority+1}"
                    return label + suffix
            return label
        return str(section + 1)

    def flags(self, index: QModelIndex) -> Qt.ItemFlag:
        if not index.isValid():
            return Qt.ItemFlag.NoItemFlags
        return (Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable |
                Qt.ItemFlag.ItemIsEditable)

    def setData(self, index: QModelIndex, value: Any,
                role: int = Qt.ItemDataRole.EditRole) -> bool:
        if role != Qt.ItemDataRole.EditRole or not index.isValid():
            return False
        r, c = index.row(), index.column()
        while len(self._data[r]) <= c:
            self._data[r].append("")
        if self._data[r][c] != str(value):
            self._data[r][c] = str(value)
            self._set_modified(True)
            self.dataChanged.emit(index, index, [role])
        return True

    # ── Ordinamento ───────────────────────────────────────────────────────────

    def sort(self, column: int, order: Qt.SortOrder = Qt.SortOrder.AscendingOrder) -> None:
        """Compatibilità Qt — chiama sort_multi con chiave singola."""
        self.sort_multi([(column, order)])

    def sort_multi(self, keys: list[tuple[int, Qt.SortOrder]]) -> None:
        """Ordinamento multi-colonna. keys = [(col, order), ...] in priorità decrescente."""
        self.layoutAboutToBeChanged.emit()
        self._sort_keys = list(keys)

        def _key(row: list[str]) -> tuple:
            parts = []
            for col, order in keys:
                val = row[col] if col < len(row) else ""
                try:
                    num = float(val.replace(",", "."))
                    parts.append((0, num if order == Qt.SortOrder.AscendingOrder else -num))
                except (ValueError, AttributeError):
                    s = val.lower()
                    parts.append((1, s if order == Qt.SortOrder.AscendingOrder else
                                  "".join(chr(0x10FFFF - ord(c)) for c in s)))
            return tuple(parts)

        self._data.sort(key=_key)
        self.layoutChanged.emit()
        # Aggiorna header per mostrare indicatori
        self.headerDataChanged.emit(Qt.Orientation.Horizontal, 0, self.columnCount() - 1)

    def clear_sort(self) -> None:
        self._sort_keys = []
        self.headerDataChanged.emit(Qt.Orientation.Horizontal, 0, self.columnCount() - 1)

    # ── Struttura ─────────────────────────────────────────────────────────────

    def add_row(self) -> None:
        self.insert_row_at(len(self._data))

    def insert_row_at(self, position: int) -> None:
        n = self.columnCount()
        position = max(0, min(position, len(self._data)))
        self.beginInsertRows(QModelIndex(), position, position)
        self._data.insert(position, [""] * n)
        self.endInsertRows()
        self._set_modified(True)

    def add_column(self, name: str = "") -> None:
        col = self.columnCount()
        self.beginInsertColumns(QModelIndex(), col, col)
        self._headers.append(name or f"Col{col + 1}")
        for row in self._data:
            row.append("")
        self.endInsertColumns()
        self._set_modified(True)

    def rename_column(self, col: int, new_name: str) -> None:
        if 0 <= col < len(self._headers) and new_name != self._headers[col]:
            self._headers[col] = new_name
            self.headerDataChanged.emit(Qt.Orientation.Horizontal, col, col)
            self._set_modified(True)

    def delete_rows(self, rows: list[int]) -> None:
        for r in sorted(set(rows), reverse=True):
            if 0 <= r < len(self._data):
                self.beginRemoveRows(QModelIndex(), r, r)
                del self._data[r]
                self.endRemoveRows()
        self._set_modified(True)

    def reorder_columns(self, new_order: list[int]) -> None:
        """new_order[i] = indice vecchio della colonna che va in posizione i."""
        self.layoutAboutToBeChanged.emit()
        self._headers = [self._headers[i] if i < len(self._headers) else "" for i in new_order]
        self._data = [
            [row[i] if i < len(row) else "" for i in new_order]
            for row in self._data
        ]
        # Aggiorna sort keys se presenti
        old_to_new = {old: new for new, old in enumerate(new_order)}
        self._sort_keys = [
            (old_to_new[c], o) for c, o in self._sort_keys if c in old_to_new
        ]
        self._set_modified(True)
        self.layoutChanged.emit()

    def reorder_rows(self, new_order: list[int]) -> None:
        """new_order[i] = indice vecchio della riga che va in posizione i."""
        self.layoutAboutToBeChanged.emit()
        self._data = [self._data[i] for i in new_order if i < len(self._data)]
        self._set_modified(True)
        self.layoutChanged.emit()

    # ── Stato ─────────────────────────────────────────────────────────────────

    def is_modified(self) -> bool:
        return self._modified

    def mark_saved(self) -> None:
        self._set_modified(False)

    def get_headers(self) -> list[str]:
        return self._headers[:]

    def get_data(self) -> list[list[str]]:
        return [row[:] for row in self._data]

    def _set_modified(self, value: bool) -> None:
        if self._modified != value:
            self._modified = value
            self.modified_changed.emit(value)


# ─── SpreadsheetIO ────────────────────────────────────────────────────────────

class SpreadsheetIO:
    EXTENSIONS = frozenset({".csv", ".tsv", ".xlsx", ".xlsm", ".xls", ".ods"})

    @staticmethod
    def get_sheet_names(path: Path) -> list[str]:
        """Restituisce i nomi dei fogli (XLSX/XLS). Lista vuota per altri formati."""
        ext = path.suffix.lower()
        try:
            if ext in (".xlsx", ".xlsm"):
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True)
                names = list(wb.sheetnames)
                wb.close()
                return names
            if ext == ".xls":
                import xlrd
                return xlrd.open_workbook(str(path)).sheet_names()
        except Exception:
            pass
        return []

    @staticmethod
    def load(path: Path, delimiter: str = ",", encoding: str = "utf-8-sig",
             first_row_header: bool = True,
             sheet: Optional[str] = None) -> tuple[list[str], list[list[str]], Optional[str]]:
        ext = path.suffix.lower()
        try:
            if ext in (".csv", ".tsv"):
                return SpreadsheetIO._load_csv(path, delimiter, encoding, first_row_header)
            elif ext in (".xlsx", ".xlsm"):
                return SpreadsheetIO._load_xlsx(path, sheet=sheet)
            elif ext == ".xls":
                return SpreadsheetIO._load_xls(path, sheet=sheet)
            elif ext == ".ods":
                return SpreadsheetIO._load_ods(path)
            else:
                return [], [], f"Formato non supportato: {ext}"
        except Exception as exc:
            return [], [], str(exc)

    @staticmethod
    def _load_csv(path: Path, delimiter: str, encoding: str,
                  first_row_header: bool) -> tuple[list[str], list[list[str]], Optional[str]]:
        with open(path, "r", encoding=encoding, newline="", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            rows = list(reader)
        if not rows:
            return [], [], None
        if first_row_header:
            headers = rows[0]
            data = rows[1:]
        else:
            n = max((len(r) for r in rows), default=1)
            headers = [f"Col{i + 1}" for i in range(n)]
            data = rows
        n = len(headers)
        data = [(row + [""] * n)[:n] for row in data]
        return headers, data, None

    @staticmethod
    def _load_xlsx(path: Path, sheet: Optional[str] = None) -> tuple[list[str], list[list[str]], Optional[str]]:
        try:
            import openpyxl
        except ImportError:
            return [], [], "openpyxl non installato.  pip install openpyxl"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        sheet_names = wb.sheetnames
        rows = [[str(c.value) if c.value is not None else "" for c in row]
                for row in ws.iter_rows()]
        wb.close()
        if not rows:
            return [], [], None
        headers = rows[0]
        data = rows[1:]
        n = len(headers)
        data = [(r + [""] * n)[:n] for r in data]
        warning = None
        if len(sheet_names) > 1:
            warning = (
                f"Il file contiene {len(sheet_names)} fogli: "
                f"{', '.join(sheet_names)}.\n"
                f"Caricato solo il foglio attivo: «{ws.title}».\n"
                f"Il supporto multi-foglio non è ancora disponibile."
            )
        return headers, data, warning

    @staticmethod
    def _load_xls(path: Path, sheet: Optional[str] = None) -> tuple[list[str], list[list[str]], Optional[str]]:
        try:
            import xlrd
        except ImportError:
            return [], [], "xlrd non installato.  pip install xlrd"
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_name(sheet) if sheet else wb.sheet_by_index(0)
        rows = [[str(ws.cell_value(r, c)) for c in range(ws.ncols)] for r in range(ws.nrows)]
        if not rows:
            return [], [], None
        headers = rows[0]
        data = rows[1:]
        n = len(headers)
        data = [(r + [""] * n)[:n] for r in data]
        warning = None
        if wb.nsheets > 1:
            names = [wb.sheet_by_index(i).name for i in range(wb.nsheets)]
            warning = (
                f"Il file contiene {wb.nsheets} fogli: "
                f"{', '.join(names)}.\n"
                f"Caricato solo il primo foglio: «{ws.name}».\n"
                f"Il supporto multi-foglio non è ancora disponibile."
            )
        return headers, data, warning

    @staticmethod
    def _load_ods(path: Path) -> tuple[list[str], list[list[str]], Optional[str]]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = [[str(c.value) if c.value is not None else "" for c in row]
                    for row in ws.iter_rows()]
            wb.close()
            if rows:
                headers = rows[0]
                data = rows[1:]
                n = len(headers)
                data = [(r + [""] * n)[:n] for r in data]
                return headers, data, None
        except Exception:
            pass
        try:
            from odf.opendocument import load as ods_load
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
            doc = ods_load(str(path))
            sheets = doc.spreadsheet.getElementsByType(Table)
            if not sheets:
                return [], [], "Nessun foglio trovato"
            rows = []
            for trow in sheets[0].getElementsByType(TableRow):
                row: list[str] = []
                for cell in trow.getElementsByType(TableCell):
                    ps = cell.getElementsByType(P)
                    row.append(str(ps[0]) if ps else "")
                rows.append(row)
            if not rows:
                return [], [], None
            headers = rows[0]
            data = rows[1:]
            n = len(headers)
            data = [(r + [""] * n)[:n] for r in data]
            return headers, data, None
        except ImportError:
            return [], [], "odfpy non installato.  pip install odfpy"

    @staticmethod
    def save(path: Path, headers: list[str], data: list[list[str]],
             delimiter: str = ",") -> Optional[str]:
        ext = path.suffix.lower()
        try:
            if ext in (".csv", ".tsv"):
                return SpreadsheetIO._save_csv(path, headers, data, delimiter)
            elif ext in (".xlsx", ".xlsm"):
                return SpreadsheetIO._save_xlsx(path, headers, data)
            elif ext == ".xls":
                new_path = path.with_suffix(".xlsx")
                err = SpreadsheetIO._save_xlsx(new_path, headers, data)
                return err or f"Salvato come XLSX: {new_path.name} (formato .xls non scrivibile)"
            elif ext == ".ods":
                return SpreadsheetIO._save_ods(path, headers, data)
            else:
                return f"Formato non supportato: {ext}"
        except Exception as exc:
            return str(exc)

    @staticmethod
    def _save_csv(path: Path, headers: list[str], data: list[list[str]],
                  delimiter: str) -> Optional[str]:
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerow(headers)
            writer.writerows(data)
        return None

    @staticmethod
    def _save_xlsx(path: Path, headers: list[str], data: list[list[str]]) -> Optional[str]:
        try:
            import openpyxl
        except ImportError:
            return "openpyxl non installato.  pip install openpyxl"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data:
            ws.append(row)
        wb.save(path)
        return None

    @staticmethod
    def _save_ods(path: Path, headers: list[str], data: list[list[str]]) -> Optional[str]:
        try:
            from odf.opendocument import OpenDocumentSpreadsheet
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
        except ImportError:
            return "odfpy non installato.  pip install odfpy"
        doc = OpenDocumentSpreadsheet()
        table = Table(name="Foglio1")
        doc.spreadsheet.addElement(table)
        for row_data in [headers] + data:
            tr = TableRow()
            table.addElement(tr)
            for val in row_data:
                tc = TableCell()
                tr.addElement(tc)
                tc.addElement(P(text=str(val)))
        doc.save(str(path))
        return None


# ─── ChartDialog ─────────────────────────────────────────────────────────────

class ChartDialog(QDialog):
    """Dialog per creare grafici (barre, linea, torta) da una selezione del foglio."""

    _TYPES = [("Barre", "bar"), ("Linea", "line"), ("Torta", "pie")]

    def __init__(self, headers: list[str], raw_data: list[list[str]], parent=None):
        super().__init__(parent)
        self._headers = headers
        self._raw_data = raw_data
        self._chart_type = "bar"
        self._fig = None
        self._canvas = None

        self.setWindowTitle("Crea grafico")
        self.setMinimumSize(640, 500)
        self.resize(720, 560)
        self._build_ui()
        self._refresh()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        type_row = QHBoxLayout()
        type_row.addWidget(QLabel("Tipo:"))
        grp = QButtonGroup(self)
        for label, key in self._TYPES:
            rb = QRadioButton(label)
            if key == "bar":
                rb.setChecked(True)
            rb.toggled.connect(lambda checked, k=key: checked and self._on_type(k))
            grp.addButton(rb)
            type_row.addWidget(rb)
        type_row.addStretch()
        layout.addLayout(type_row)

        try:
            from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
            from matplotlib.figure import Figure
            self._fig = Figure(figsize=(7, 4), tight_layout=True)
            self._canvas = FigureCanvas(self._fig)
            layout.addWidget(self._canvas, 1)
        except ImportError:
            lbl = QLabel("matplotlib non disponibile.\nInstalla: pip install matplotlib")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(lbl, 1)

        btn_row = QHBoxLayout()
        btn_save = QPushButton("💾 Salva immagine…")
        btn_save.clicked.connect(self._save_image)
        btn_row.addWidget(btn_save)
        btn_row.addStretch()
        btn_close = QPushButton("Chiudi")
        btn_close.setDefault(True)
        btn_close.clicked.connect(self.accept)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

    def _on_type(self, key: str) -> None:
        self._chart_type = key
        self._refresh()

    def _refresh(self) -> None:
        if self._fig is None or self._canvas is None:
            return
        self._fig.clear()
        ax = self._fig.add_subplot(111)

        headers = self._headers
        raw = self._raw_data

        if not raw:
            ax.text(0.5, 0.5, "Nessun dato", ha="center", va="center",
                    transform=ax.transAxes)
            self._canvas.draw()
            return

        # Prima colonna come etichette se non numerica, altrimenti indici
        first_numeric = True
        for row in raw[:5]:
            try:
                float((row[0] if row else "").replace(",", "."))
            except (ValueError, AttributeError):
                first_numeric = False
                break

        if not first_numeric and len(headers) > 1:
            labels = [row[0] if row else "" for row in raw]
            data_cols = list(range(1, len(headers)))
        else:
            labels = [str(i + 1) for i in range(len(raw))]
            data_cols = list(range(len(headers)))

        series: dict[str, list[float]] = {}
        for c in data_cols:
            col_name = headers[c] if c < len(headers) else f"Col{c + 1}"
            vals = []
            for row in raw:
                val = row[c] if c < len(row) else ""
                try:
                    vals.append(float(val.replace(",", ".")))
                except (ValueError, AttributeError):
                    vals.append(0.0)
            series[col_name] = vals

        if not series:
            ax.text(0.5, 0.5, "Nessun dato numerico nella selezione",
                    ha="center", va="center", transform=ax.transAxes)
            self._canvas.draw()
            return

        display_labels = [str(l)[:20] for l in labels]

        if self._chart_type == "pie":
            name, vals = next(iter(series.items()))
            if len(vals) > 12:
                vals, display_labels = vals[:12], display_labels[:12]
            abs_vals = [abs(v) for v in vals]
            if sum(abs_vals) == 0:
                ax.text(0.5, 0.5, "Valori tutti zero", ha="center", va="center",
                        transform=ax.transAxes)
            else:
                ax.pie(abs_vals, labels=display_labels, autopct="%1.1f%%", startangle=90)
                ax.set_title(name)

        elif self._chart_type == "bar":
            n_series = len(series)
            width = 0.8 / max(n_series, 1)
            x = list(range(len(labels)))
            for i, (name, vals) in enumerate(series.items()):
                offsets = [xi + (i - (n_series - 1) / 2) * width for xi in x]
                ax.bar(offsets, vals, width, label=name)
            ax.set_xticks(x)
            ax.set_xticklabels(display_labels, rotation=45, ha="right", fontsize=8)
            if n_series > 1:
                ax.legend(fontsize=8)

        else:  # line
            x = list(range(len(labels)))
            for name, vals in series.items():
                ax.plot(x, vals, marker="o", markersize=3, label=name)
            ax.set_xticks(x)
            ax.set_xticklabels(display_labels, rotation=45, ha="right", fontsize=8)
            if len(series) > 1:
                ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        self._canvas.draw()

    def _save_image(self) -> None:
        if self._fig is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Salva grafico", "", "PNG (*.png);;SVG (*.svg);;PDF (*.pdf)"
        )
        if path:
            self._fig.savefig(path, dpi=150, bbox_inches="tight")


# ─── SpreadsheetWidget ────────────────────────────────────────────────────────

class SpreadsheetWidget(QWidget):
    """Widget principale foglio di calcolo. Va inserito come tab nel TabManager."""

    modified_changed = pyqtSignal(bool)
    convert_to_text  = pyqtSignal(str, str)   # (content, suggested_filename)

    def __init__(self, path: Path, headers: list[str], data: list[list[str]],
                 read_only: bool = False, delimiter: str = ",",
                 encoding: str = "utf-8-sig", first_row_header: bool = True,
                 sheet_names: Optional[list[str]] = None, current_sheet: str = "",
                 parent=None):
        super().__init__(parent)
        self.file_path = path
        self._read_only = read_only
        self._delimiter = delimiter
        self._encoding = encoding
        self._first_row_header = first_row_header
        self._sort_keys: list[tuple[int, Qt.SortOrder]] = []
        self._sheet_names: list[str] = sheet_names or []
        self._current_sheet: str = current_sheet

        self._model = SpreadsheetModel(headers, data)
        self._model.modified_changed.connect(self._on_model_modified)

        self._proxy = _FilterProxy(self)
        self._proxy.setSourceModel(self._model)

        self._build_ui()
        self._sheet_bar_widget = self._build_sheet_bar(self._sheet_names, self._current_sheet)
        if self._sheet_bar_widget is not None:
            self.layout().addWidget(self._sheet_bar_widget)

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)

        # ── Toolbar ───────────────────────────────────────────────────────────
        toolbar = QHBoxLayout()

        self._file_label = QLabel(self.file_path.name)
        self._file_label.setStyleSheet("font-weight: bold; font-size: 12px;")
        toolbar.addWidget(self._file_label)

        if self._read_only:
            ro_lbl = QLabel("⚠ Sola lettura (.xls)")
            ro_lbl.setStyleSheet("color: orange; font-size: 11px; margin-left: 12px;")
            toolbar.addWidget(ro_lbl)

        toolbar.addStretch()

        if not self._read_only:
            btn_ins_row = QPushButton("+ Riga")
            btn_ins_row.setFixedHeight(24)
            btn_ins_row.setToolTip(
                "Inserisce una riga vuota sotto l'ultima riga selezionata\n"
                "(in fondo se nessuna riga è selezionata)"
            )
            btn_ins_row.clicked.connect(self._add_row_below)
            toolbar.addWidget(btn_ins_row)

            btn_add_col = QPushButton("+ Colonna")
            btn_add_col.setFixedHeight(24)
            btn_add_col.setToolTip("Aggiunge una colonna vuota a destra")
            btn_add_col.clicked.connect(self._model.add_column)
            toolbar.addWidget(btn_add_col)

            btn_del_rows = QPushButton("− Righe sel.")
            btn_del_rows.setFixedHeight(24)
            btn_del_rows.setToolTip("Elimina le righe selezionate")
            btn_del_rows.clicked.connect(self._delete_selected_rows)
            toolbar.addWidget(btn_del_rows)

        btn_filter = QPushButton("🔍 Filtro")
        btn_filter.setFixedHeight(24)
        btn_filter.setCheckable(True)
        btn_filter.setToolTip("Mostra/nasconde la barra filtri (Ctrl+F)")
        btn_filter.clicked.connect(self._toggle_filter_bar)
        toolbar.addWidget(btn_filter)
        self._btn_filter = btn_filter

        if not self._read_only:
            self._btn_save = QPushButton("💾 Salva")
            self._btn_save.setFixedHeight(24)
            self._btn_save.setEnabled(False)
            self._btn_save.setToolTip("Salva il foglio sul disco (Ctrl+S)")
            self._btn_save.clicked.connect(self._save)
            toolbar.addWidget(self._btn_save)

        btn_save_as = QPushButton("📤 Esporta/Salva come…")
        btn_save_as.setFixedHeight(24)
        btn_save_as.setToolTip(
            "Salva il foglio in un nuovo file scegliendo il formato\n"
            "Conversione disponibile: CSV ↔ XLSX ↔ ODS ↔ TSV"
        )
        btn_save_as.clicked.connect(self._save_as)
        toolbar.addWidget(btn_save_as)

        btn_md = QPushButton("→ Markdown")
        btn_md.setFixedHeight(24)
        btn_md.setToolTip(
            "Converte il foglio in tabella Markdown\n"
            "Apre il risultato in una nuova scheda"
        )
        btn_md.clicked.connect(self._convert_markdown)
        toolbar.addWidget(btn_md)

        btn_tex = QPushButton("→ tabularx")
        btn_tex.setFixedHeight(24)
        btn_tex.setToolTip(
            "Converte il foglio in ambiente tabularx LaTeX\n"
            "Apre il risultato in una nuova scheda"
        )
        btn_tex.clicked.connect(self._convert_tabularx)
        toolbar.addWidget(btn_tex)

        btn_chart = QPushButton("📊 Grafico")
        btn_chart.setFixedHeight(24)
        btn_chart.setToolTip(
            "Crea un grafico dalla selezione corrente\n"
            "\n"
            "Come usarlo:\n"
            "  1. Seleziona le celle che vuoi visualizzare\n"
            "  2. Clicca questo pulsante\n"
            "  3. Scegli il tipo: Barre, Linea o Torta\n"
            "\n"
            "Formato dati atteso:\n"
            "  • Prima colonna = etichette (es. nomi, mesi) — se non numerica\n"
            "  • Colonne successive = serie numeriche da tracciare\n"
            "  • Se tutte le colonne sono numeriche, l'asse X usa i numeri di riga\n"
            "\n"
            "Note:\n"
            "  • La torta usa solo la prima serie numerica (max 12 valori)\n"
            "  • Il grafico può essere salvato come PNG, SVG o PDF"
        )
        btn_chart.clicked.connect(self._create_chart)
        toolbar.addWidget(btn_chart)

        layout.addLayout(toolbar)

        # ── Barra formula (fx) ────────────────────────────────────────────────
        self._formula_bar_updating = False
        self._formula_inserting = False
        self._formula_bar_active = False   # True mentre la formula bar ha o aveva focus
        self._formula_focus_timer = QTimer()
        self._formula_focus_timer.setSingleShot(True)
        self._formula_focus_timer.setInterval(150)
        self._formula_focus_timer.timeout.connect(
            lambda: setattr(self, '_formula_bar_active', False)
        )
        fx_row = QHBoxLayout()
        fx_row.setContentsMargins(0, 0, 0, 0)
        fx_row.setSpacing(4)
        self._fx_btn = QPushButton("fx ▾")
        self._fx_btn.setFixedSize(44, 24)
        self._fx_btn.setFlat(True)
        self._fx_btn.setStyleSheet(
            "QPushButton { font-weight: bold; font-size: 12px; }"
            "QPushButton:hover { background: #3a3a5a; border-radius: 3px; }"
        )
        self._fx_btn.setToolTip("Inserisci funzione…\nMostra l'elenco delle funzioni disponibili")
        self._fx_btn.clicked.connect(self._show_formula_menu)
        fx_row.addWidget(self._fx_btn)
        self._cell_ref = QLineEdit()
        self._cell_ref.setReadOnly(True)
        self._cell_ref.setFixedWidth(52)
        self._cell_ref.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._cell_ref.setPlaceholderText("A1")
        self._cell_ref.setToolTip("Indirizzo della cella selezionata")
        fx_row.addWidget(self._cell_ref)
        self._formula_bar = QLineEdit()
        self._formula_bar.setPlaceholderText("Seleziona una cella per modificarla…")
        self._formula_bar.setToolTip(
            "Mostra e modifica il contenuto grezzo della cella selezionata.\n"
            "Le formule iniziano con = (esempi):\n"
            "  =SUM(A1:A10)        somma da A1 ad A10\n"
            "  =AVERAGE(B1:B5)     media\n"
            "  =IF(A1>0,\"sì\",\"no\")  condizione\n"
            "  =A1+B1*2            aritmetica\n"
            "  =CONCAT(A1,\" \",B1)  testo\n"
            "\nPremi Invio per confermare, Esc per annullare."
        )
        self._formula_bar.returnPressed.connect(self._apply_formula_bar)
        self._formula_bar.installEventFilter(self)
        fx_row.addWidget(self._formula_bar)
        layout.addLayout(fx_row)

        # ── Barra filtri (nascosta di default) ────────────────────────────────
        self._filter_bar = self._build_filter_bar()
        self._filter_bar.hide()
        layout.addWidget(self._filter_bar)

        # ── Tabella ───────────────────────────────────────────────────────────
        self._view = QTableView()
        self._view.setModel(self._proxy)
        self._view.setSortingEnabled(False)   # sort manuale via header click
        self._view.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._view.setAlternatingRowColors(True)

        hh = self._view.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hh.setStretchLastSection(True)
        hh.setSortIndicatorShown(True)
        hh.setSectionsMovable(True)
        hh.sectionClicked.connect(self._on_header_clicked)
        hh.sectionMoved.connect(self._on_column_moved)
        hh.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        hh.customContextMenuRequested.connect(self._on_header_context_menu)
        hh.setToolTip(
            "Click: ordina per questa colonna\n"
            "Shift+Click: aggiunge colonna all'ordinamento\n"
            "Trascina: sposta la colonna\n"
            "Tasto destro: rinomina colonna"
        )

        vh = self._view.verticalHeader()
        vh.setDefaultSectionSize(22)
        vh.setSectionsMovable(True)
        vh.setToolTip("Trascina per spostare la riga")
        vh.sectionMoved.connect(self._on_row_moved)

        if self._read_only:
            self._view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        else:
            self._view.setEditTriggers(
                QAbstractItemView.EditTrigger.DoubleClicked |
                QAbstractItemView.EditTrigger.EditKeyPressed |
                QAbstractItemView.EditTrigger.AnyKeyPressed
            )

        self._view.resizeColumnsToContents()
        self._view.selectionModel().selectionChanged.connect(self._on_selection_changed)
        self._view.selectionModel().currentChanged.connect(self._on_current_cell_changed)
        # Intercetta Ctrl+S e Ctrl+F direttamente sul viewport della tabella
        self._view.viewport().installEventFilter(self)
        layout.addWidget(self._view, 1)

        # ── Status bar ────────────────────────────────────────────────────────
        self._status = QLabel("Pronto")
        self._status.setStyleSheet("font-size: 11px; color: #888;")
        layout.addWidget(self._status)

    def _build_filter_bar(self) -> QWidget:
        """Barra filtro testo (aperta da pulsante 🔍 Filtro o Ctrl+F)."""
        bar = QWidget()
        bar.setStyleSheet("background: #2a2a3a; border-radius: 4px; padding: 2px;")
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(6, 2, 6, 2)
        lay.setSpacing(6)

        lay.addWidget(QLabel("🔍 Filtra:"))

        self._filter_col = QComboBox()
        self._filter_col.setFixedHeight(22)
        self._filter_col.setMinimumWidth(120)
        self._filter_col.addItem("— Tutte le colonne —", -1)
        for i, h in enumerate(self._model.get_headers()):
            self._filter_col.addItem(h, i)
        self._filter_col.currentIndexChanged.connect(lambda _: self._apply_filter())
        lay.addWidget(self._filter_col)

        self._filter_text = QLineEdit()
        self._filter_text.setFixedHeight(22)
        self._filter_text.setPlaceholderText("testo da cercare…")
        self._filter_text.setMinimumWidth(180)
        self._filter_text.setStyleSheet(
            "QLineEdit { background: #1e1e2e; color: #dddddd;"
            " border: 1px solid #555; border-radius: 2px; padding: 1px 4px; }"
        )
        self._filter_text.textChanged.connect(self._apply_filter)
        lay.addWidget(self._filter_text)

        self._filter_active_label = QLabel("")
        self._filter_active_label.setStyleSheet("color: #f0a030; font-size: 11px;")
        lay.addWidget(self._filter_active_label)

        lay.addStretch()

        btn_clear = QPushButton("✗ Pulisci")
        btn_clear.setFixedHeight(22)
        btn_clear.setToolTip("Rimuove il filtro attivo")
        btn_clear.clicked.connect(self._clear_filters)
        lay.addWidget(btn_clear)

        return bar

    # ── Slot toolbar / filtri ─────────────────────────────────────────────────

    def _toggle_filter_bar(self, checked: bool) -> None:
        if checked:
            self._filter_bar.show()
            from PyQt6.QtCore import QTimer
            QTimer.singleShot(0, self._filter_text.setFocus)
        else:
            self._filter_bar.hide()
            self._proxy.clear()
            self._update_filter_label()

    def _apply_filter(self) -> None:
        text = self._filter_text.text()
        col = self._filter_col.currentData()
        if text.strip():
            self._proxy.set_filter(col, text)
        else:
            self._proxy.clear()
        self._update_filter_label()

    def _clear_filters(self) -> None:
        self._filter_text.clear()
        self._proxy.clear()
        self._update_filter_label()

    def _update_filter_label(self) -> None:
        text = self._proxy.active_text()
        if text:
            col = self._proxy.active_col()
            headers = self._model.get_headers()
            col_name = ("Tutte" if col == -1
                        else headers[col] if col < len(headers) else str(col))
            n_vis = self._proxy.rowCount()
            n_tot = self._model.rowCount()
            self._filter_active_label.setText(
                f"Filtro: {col_name} «{text}»  ({n_vis}/{n_tot} righe)"
            )
        else:
            self._filter_active_label.setText("")

    # ── Event filter (Ctrl+S, Ctrl+F) ────────────────────────────────────────

    def eventFilter(self, obj, event) -> bool:
        from PyQt6.QtCore import QEvent
        if obj is self._formula_bar:
            if event.type() == QEvent.Type.FocusIn:
                self._formula_bar_active = True
                self._formula_focus_timer.stop()
                return False
            if event.type() == QEvent.Type.FocusOut:
                # Ritardo: il MouseButtonPress sul viewport arriva dopo FocusOut
                self._formula_focus_timer.start()
                return False
        if obj is self._formula_bar and event.type() == QEvent.Type.KeyPress:
            if event.key() == Qt.Key.Key_Escape:
                # Ripristina il valore originale e sposta il focus alla tabella
                current = self._view.selectionModel().currentIndex()
                if current.isValid():
                    src = self._proxy.mapToSource(current)
                    self._formula_bar_updating = True
                    self._formula_bar.setText(
                        str(self._model.data(src, Qt.ItemDataRole.EditRole) or "")
                    )
                    self._formula_bar_updating = False
                self._formula_bar_active = False
                self._formula_focus_timer.stop()
                self._view.setFocus()
                return True
        if hasattr(self, '_view') and obj is self._view.viewport():
            if (event.type() == QEvent.Type.MouseButtonPress
                    and self._formula_bar_active
                    and self._formula_bar.text().startswith("=")):
                idx = self._view.indexAt(event.pos())
                if idx.isValid():
                    src = self._proxy.mapToSource(idx)
                    ref = f"{_col_letter(src.column())}{src.row() + 1}"
                    cp = self._formula_bar.cursorPosition()
                    text = self._formula_bar.text()
                    self._formula_bar.setText(text[:cp] + ref + text[cp:])
                    self._formula_bar.setCursorPosition(cp + len(ref))
                    self._formula_bar.setFocus()
                    return True  # consuma: la selezione rimane sulla cella originale
            if event.type() == QEvent.Type.KeyPress:
                mods = event.modifiers()
                ctrl = Qt.KeyboardModifier.ControlModifier
                if mods == ctrl:
                    if event.key() == Qt.Key.Key_S:
                        if not self._read_only:
                            self._save()
                        return True
                    if event.key() == Qt.Key.Key_F:
                        self._btn_filter.setChecked(True)
                        self._toggle_filter_bar(True)
                        return True
        return super().eventFilter(obj, event)

    # ── Ordinamento multi-colonna ─────────────────────────────────────────────

    def _on_header_clicked(self, col: int) -> None:
        from PyQt6.QtWidgets import QApplication
        modifiers = QApplication.keyboardModifiers()

        if modifiers & Qt.KeyboardModifier.ShiftModifier:
            # Aggiunge/toglie questa colonna dall'ordinamento
            existing = [(c, o) for c, o in self._sort_keys if c == col]
            if existing:
                # Inverte l'ordine
                new_order = Qt.SortOrder.DescendingOrder if existing[0][1] == Qt.SortOrder.AscendingOrder else Qt.SortOrder.AscendingOrder
                self._sort_keys = [(c, new_order if c == col else o)
                                   for c, o in self._sort_keys]
            else:
                self._sort_keys.append((col, Qt.SortOrder.AscendingOrder))
        else:
            # Ordinamento singolo: primo click ASC, secondo DESC, terzo rimuove
            if len(self._sort_keys) == 1 and self._sort_keys[0][0] == col:
                if self._sort_keys[0][1] == Qt.SortOrder.AscendingOrder:
                    self._sort_keys = [(col, Qt.SortOrder.DescendingOrder)]
                else:
                    self._sort_keys = []
                    self._model.clear_sort()
                    self._view.horizontalHeader().setSortIndicator(-1, Qt.SortOrder.AscendingOrder)
                    return
            else:
                self._sort_keys = [(col, Qt.SortOrder.AscendingOrder)]

        if self._sort_keys:
            self._model.sort_multi(self._sort_keys)
            primary_col, primary_order = self._sort_keys[0]
            self._view.horizontalHeader().setSortIndicator(primary_col, primary_order)

    # ── Spostamento colonne/righe (drag header) ────────────────────────────────

    def _on_column_moved(self, logical: int, old_visual: int, new_visual: int) -> None:
        hh = self._view.horizontalHeader()
        n = self._model.columnCount()
        new_order = [hh.logicalIndex(i) for i in range(n)]
        hh.blockSignals(True)
        self._model.reorder_columns(new_order)
        for visual in range(n):
            cur = hh.visualIndex(visual)
            if cur != visual:
                hh.moveSection(cur, visual)
        hh.blockSignals(False)
        self._sort_keys = list(self._model._sort_keys)
        # Filtre testo rimane valido (usa indici logici che non cambiano)
        self._update_filter_label()

    def _on_row_moved(self, logical: int, old_visual: int, new_visual: int) -> None:
        vh = self._view.verticalHeader()
        n = self._model.rowCount()
        new_order = [vh.logicalIndex(i) for i in range(n)]
        vh.blockSignals(True)
        self._model.reorder_rows(new_order)
        for visual in range(n):
            cur = vh.visualIndex(visual)
            if cur != visual:
                vh.moveSection(cur, visual)
        vh.blockSignals(False)

    # ── Righe ─────────────────────────────────────────────────────────────────

    def _add_row_below(self) -> None:
        sel = self._view.selectionModel().selectedIndexes()
        if sel:
            max_proxy_row = max(idx.row() for idx in sel)
            src_idx = self._proxy.mapToSource(self._proxy.index(max_proxy_row, 0))
            insert_pos = src_idx.row() + 1
        else:
            insert_pos = self._model.rowCount()
        self._model.insert_row_at(insert_pos)

    def _delete_selected_rows(self) -> None:
        sel = self._view.selectionModel().selectedIndexes()
        if not sel:
            return
        proxy_rows = {idx.row() for idx in sel}
        source_rows = []
        for pr in proxy_rows:
            src = self._proxy.mapToSource(self._proxy.index(pr, 0))
            source_rows.append(src.row())
        self._model.delete_rows(source_rows)

    # ── Header context menu / Rinomina colonna ────────────────────────────────

    def _on_header_context_menu(self, pos) -> None:
        if self._read_only:
            return
        hh = self._view.horizontalHeader()
        logical = hh.logicalIndexAt(pos)
        if logical < 0:
            return
        menu = QMenu(self)
        act = menu.addAction("Rinomina colonna…")
        if menu.exec(hh.mapToGlobal(pos)) == act:
            headers = self._model.get_headers()
            current = headers[logical] if logical < len(headers) else ""
            QTimer.singleShot(0, lambda: self._do_rename_column(logical, current))

    def _do_rename_column(self, col: int, current: str) -> None:
        new_name, ok = QInputDialog.getText(
            self, "Rinomina colonna", "Nuovo nome:", text=current
        )
        if ok and new_name.strip() and new_name.strip() != current:
            self._model.rename_column(col, new_name.strip())

    def _create_chart(self) -> None:
        sel = self._view.selectionModel().selectedIndexes()
        if not sel:
            QMessageBox.information(
                self, "Grafico", "Seleziona almeno una colonna di dati per creare un grafico."
            )
            return

        proxy_rows = sorted({idx.row() for idx in sel})
        proxy_cols = sorted({idx.column() for idx in sel})

        src_cols = [
            self._proxy.mapToSource(self._proxy.index(0, c)).column()
            for c in proxy_cols
        ]
        all_headers = self._model.get_headers()
        headers = [all_headers[c] if c < len(all_headers) else str(c + 1) for c in src_cols]

        raw_data: list[list[str]] = []
        for pr in proxy_rows:
            row = []
            for pc in proxy_cols:
                idx = self._proxy.index(pr, pc)
                row.append(self._proxy.data(idx, Qt.ItemDataRole.DisplayRole) or "")
            raw_data.append(row)

        dlg = ChartDialog(headers, raw_data, self)
        dlg.exec()

    # ── Formula bar ───────────────────────────────────────────────────────────

    def _on_current_cell_changed(self, current: QModelIndex, _prev: QModelIndex) -> None:
        if self._formula_inserting:
            self._formula_inserting = False
            return  # la formula bar è già aggiornata, non sovrascrivere
        if not current.isValid():
            self._formula_bar.setText("")
            self._cell_ref.setText("")
            return
        src = self._proxy.mapToSource(current)
        r, c = src.row(), src.column()
        self._cell_ref.setText(f"{_col_letter(c)}{r + 1}")
        self._formula_bar_updating = True
        self._formula_bar.setText(str(self._model.data(src, Qt.ItemDataRole.EditRole) or ""))
        self._formula_bar_updating = False

    def _apply_formula_bar(self) -> None:
        if self._formula_bar_updating or self._read_only:
            return
        current = self._view.selectionModel().currentIndex()
        if not current.isValid():
            return
        src = self._proxy.mapToSource(current)
        self._model.setData(src, self._formula_bar.text(), Qt.ItemDataRole.EditRole)
        self._formula_bar_active = False
        self._formula_focus_timer.stop()
        self._view.setFocus()

    def _show_formula_menu(self) -> None:
        menu = QMenu(self)
        menu.setToolTipsVisible(True)
        for category, funcs in _FORMULA_MENU:
            sub = menu.addMenu(category)
            sub.setToolTipsVisible(True)
            for label, template, tooltip in funcs:
                act = sub.addAction(label)
                act.setToolTip(tooltip)
                act.triggered.connect(
                    lambda checked, t=template: self._insert_formula_template(t)
                )
        global_pos = self._fx_btn.mapToGlobal(
            self._fx_btn.rect().bottomLeft()
        )
        menu.exec(global_pos)

    def _insert_formula_template(self, template: str) -> None:
        current = self._view.selectionModel().currentIndex()
        if not current.isValid():
            return
        cp = self._formula_bar.cursorPosition()
        text = self._formula_bar.text()
        # Se la barra è vuota parti da zero, altrimenti inserisci al cursore
        if not text:
            new_text = template
            new_cp = len(template)
        else:
            new_text = text[:cp] + template + text[cp:]
            new_cp = cp + len(template)
        self._formula_bar.setText(new_text)
        self._formula_bar.setCursorPosition(new_cp)
        self._formula_bar.setFocus()

    # ── Barra fogli ───────────────────────────────────────────────────────────

    def _build_sheet_bar(self, sheet_names: list[str], current: str) -> Optional[QWidget]:
        if len(sheet_names) <= 1:
            return None
        bar = QWidget()
        bar.setStyleSheet("background: #1e1e2e; border-top: 1px solid #444;")
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(4, 2, 4, 2)
        lay.setSpacing(2)
        self._sheet_buttons: dict[str, QPushButton] = {}
        for name in sheet_names:
            btn = QPushButton(name)
            btn.setFixedHeight(22)
            btn.setCheckable(True)
            btn.setChecked(name == current)
            self._update_sheet_btn_style(btn, name == current)
            btn.clicked.connect(lambda _checked, n=name: self._load_sheet(n))
            self._sheet_buttons[name] = btn
            lay.addWidget(btn)
        lay.addStretch()
        return bar

    def _update_sheet_btn_style(self, btn: QPushButton, active: bool) -> None:
        if active:
            btn.setStyleSheet(
                "QPushButton { background:#3a3a5a; color:#ffffff;"
                " border:1px solid #7070c0; border-radius:2px; padding:0 8px; }"
            )
        else:
            btn.setStyleSheet(
                "QPushButton { background:#2a2a3a; color:#aaaaaa;"
                " border:1px solid #444; border-radius:2px; padding:0 8px; }"
                "QPushButton:hover { background:#333355; color:#dddddd; }"
            )

    def _load_sheet(self, sheet_name: str) -> None:
        if sheet_name == self._current_sheet:
            return
        headers, data, error = SpreadsheetIO.load(self.file_path, sheet=sheet_name)
        if error and not headers:
            QMessageBox.warning(self, "Errore caricamento foglio", error)
            return
        if not headers and not data:
            headers = [f"Col{i+1}" for i in range(5)]
            data = [[""] * 5 for _ in range(20)]
        self._model.modified_changed.disconnect(self._on_model_modified)
        self._model = SpreadsheetModel(headers, data)
        self._model.modified_changed.connect(self._on_model_modified)
        self._proxy = _FilterProxy(self)
        self._proxy.setSourceModel(self._model)
        self._view.setModel(self._proxy)
        self._view.selectionModel().selectionChanged.connect(self._on_selection_changed)
        self._view.selectionModel().currentChanged.connect(self._on_current_cell_changed)
        self._sort_keys = []
        self._current_sheet = sheet_name
        self._formula_bar.setText("")
        self._cell_ref.setText("")
        self._status.setText("Pronto")
        if hasattr(self, "_sheet_buttons"):
            for name, btn in self._sheet_buttons.items():
                btn.setChecked(name == sheet_name)
                self._update_sheet_btn_style(btn, name == sheet_name)

    # ── Selezione / Status ────────────────────────────────────────────────────

    def _on_selection_changed(self) -> None:
        sel = self._view.selectionModel().selectedIndexes()
        if not sel:
            self._status.setText("Pronto")
            return
        rows = {idx.row() for idx in sel}
        cols = {idx.column() for idx in sel}
        nums: list[float] = []
        for idx in sel:
            src = self._proxy.mapToSource(idx)
            v = self._model.data(src, Qt.ItemDataRole.DisplayRole)
            if v:
                try:
                    nums.append(float(v.replace(",", ".")))
                except (ValueError, AttributeError):
                    pass
        parts = [f"Selezione: {len(rows)}r × {len(cols)}c ({len(sel)} celle)"]
        if nums:
            parts += [
                f"Somma: {sum(nums):.6g}",
                f"Media: {sum(nums)/len(nums):.6g}",
                f"Min: {min(nums):.6g}",
                f"Max: {max(nums):.6g}",
                f"Num: {len(nums)}",
            ]
        self._status.setText("   |   ".join(parts))

    # ── Salvataggio ───────────────────────────────────────────────────────────

    def _on_model_modified(self, modified: bool) -> None:
        if not self._read_only:
            self._btn_save.setEnabled(modified)
        self.modified_changed.emit(modified)

    def _save(self) -> None:
        if self._read_only:
            # XLS sola lettura: offri salva-come
            self._save_as()
            return
        err = SpreadsheetIO.save(
            self.file_path,
            self._model.get_headers(),
            self._model.get_data(),
            self._delimiter,
        )
        if err:
            QMessageBox.warning(self, "Errore salvataggio", err)
        else:
            self._model.mark_saved()

    def _save_as(self) -> None:
        _FILTERS = [
            ("CSV (*.csv)",            ".csv"),
            ("Excel XLSX (*.xlsx)",    ".xlsx"),
            ("ODS LibreOffice (*.ods)",".ods"),
            ("TSV (*.tsv)",            ".tsv"),
        ]
        filter_strings = [f for f, _ in _FILTERS]

        ext = self.file_path.suffix.lower()
        initial_filter = filter_strings[0]
        for fname, fext in _FILTERS:
            if fext == ext:
                initial_filter = fname
                break

        def _ext_for_filter(f: str) -> str:
            for fname, fext in _FILTERS:
                if fname == f:
                    return fext
            return ".csv"

        dlg = QFileDialog(self, "Salva foglio come…")
        dlg.setAcceptMode(QFileDialog.AcceptMode.AcceptSave)
        dlg.setDirectory(str(self.file_path.parent))
        dlg.setNameFilters(filter_strings)
        dlg.selectNameFilter(initial_filter)
        dlg.setDefaultSuffix(ext.lstrip(".") or "csv")
        dlg.selectFile(self.file_path.stem)

        def _on_filter_changed(f: str) -> None:
            new_ext = _ext_for_filter(f)
            dlg.setDefaultSuffix(new_ext.lstrip("."))
            files = dlg.selectedFiles()
            if files:
                dlg.selectFile(Path(files[0]).with_suffix(new_ext).name)

        dlg.filterSelected.connect(_on_filter_changed)

        if dlg.exec() != QFileDialog.DialogCode.Accepted:
            return
        files = dlg.selectedFiles()
        if not files:
            return

        dest_path = Path(files[0])
        # Garantisce l'estensione corretta in base al filtro scelto
        expected_ext = _ext_for_filter(dlg.selectedNameFilter())
        if dest_path.suffix.lower() != expected_ext:
            dest_path = dest_path.with_suffix(expected_ext)

        delim = "\t" if dest_path.suffix.lower() == ".tsv" else self._delimiter
        err = SpreadsheetIO.save(
            dest_path, self._model.get_headers(), self._model.get_data(), delim
        )
        if err:
            QMessageBox.warning(self, "Errore salvataggio", err)
        else:
            self.file_path = dest_path
            self._delimiter = delim
            self._read_only = False
            self._file_label.setText(dest_path.name)
            self._model.mark_saved()

    # ── Conversione testo ─────────────────────────────────────────────────────

    @staticmethod
    def _build_markdown_table(headers: list[str], data: list[list[str]]) -> str:
        def esc(s: str) -> str:
            return str(s).replace("|", "\\|").replace("\n", " ")

        widths = [max(3, len(esc(h))) for h in headers]
        for row in data:
            for i in range(len(headers)):
                cell = esc(row[i]) if i < len(row) else ""
                widths[i] = max(widths[i], len(cell))

        def pad(s: str, w: int) -> str:
            return esc(s).ljust(w)

        header_line = "| " + " | ".join(pad(h, widths[i]) for i, h in enumerate(headers)) + " |"
        sep_line = "|" + "|".join("-" * (w + 2) for w in widths) + "|"
        rows = []
        for row in data:
            cells = [pad(row[i] if i < len(row) else "", widths[i]) for i in range(len(headers))]
            rows.append("| " + " | ".join(cells) + " |")
        return "\n".join([header_line, sep_line] + rows)

    @staticmethod
    def _build_tabularx_table(headers: list[str], data: list[list[str]],
                               caption: str = "") -> str:
        _TEX_SPECIAL = str.maketrans({
            "&": "\\&", "%": "\\%", "$": "\\$", "#": "\\#",
            "_": "\\_", "{": "\\{", "}": "\\}", "~": "\\textasciitilde{}",
            "^": "\\textasciicircum{}", "\\": "\\textbackslash{}",
        })

        def esc(s: str) -> str:
            return str(s).translate(_TEX_SPECIAL)

        n = len(headers)
        col_spec = "X" * n if n else "X"
        header_row = " & ".join(f"\\textbf{{{esc(h)}}}" for h in headers) + " \\\\"
        rows_lines = [" & ".join(esc(row[i] if i < len(row) else "") for i in range(n)) + " \\\\"
                      for row in data]
        return "\n".join([
            "% Richiede: \\usepackage{tabularx,booktabs}",
            "\\begin{table}[htbp]",
            "\\centering",
            f"\\begin{{tabularx}}{{\\textwidth}}{{{col_spec}}}",
            "\\toprule",
            header_row,
            "\\midrule",
            *rows_lines,
            "\\bottomrule",
            "\\end{tabularx}",
            f"\\caption{{{esc(caption)}}}",
            "\\label{tab:}",
            "\\end{table}",
        ])

    def _ask_sheets_scope(self, title: str) -> str:
        """Chiede all'utente se convertire il foglio corrente o tutti i fogli.
        Restituisce 'current', 'all', o 'cancel'."""
        from PyQt6.QtWidgets import QMessageBox
        msg = QMessageBox(self)
        msg.setWindowTitle(title)
        msg.setText(
            f"Il file contiene {len(self._sheet_names)} fogli.\n"
            "Vuoi convertire solo il foglio corrente o tutti i fogli?"
        )
        btn_curr = msg.addButton("Solo foglio corrente", QMessageBox.ButtonRole.AcceptRole)
        btn_all  = msg.addButton("Tutti i fogli",        QMessageBox.ButtonRole.ActionRole)
        btn_cancel = msg.addButton("Annulla",            QMessageBox.ButtonRole.RejectRole)
        msg.exec()
        clicked = msg.clickedButton()
        if clicked is btn_all:
            return "all"
        if clicked is btn_curr:
            return "current"
        return "cancel"

    def _convert_markdown(self) -> None:
        scope = "current"
        if len(self._sheet_names) > 1:
            scope = self._ask_sheets_scope("Converti in Markdown")
            if scope == "cancel":
                return

        if scope == "all":
            parts: list[str] = []
            for sheet_name in self._sheet_names:
                headers, data, _ = SpreadsheetIO.load(self.file_path, sheet=sheet_name)
                parts.append(f"## {sheet_name}\n\n" + self._build_markdown_table(headers, data))
            content = "\n\n---\n\n".join(parts)
        else:
            content = self._build_markdown_table(
                self._model.get_headers(), self._model.get_data()
            )

        self.convert_to_text.emit(content, f"{self.file_path.stem}.md")

    def _convert_tabularx(self) -> None:
        scope = "current"
        if len(self._sheet_names) > 1:
            scope = self._ask_sheets_scope("Converti in tabularx")
            if scope == "cancel":
                return

        if scope == "all":
            parts: list[str] = []
            for sheet_name in self._sheet_names:
                headers, data, _ = SpreadsheetIO.load(self.file_path, sheet=sheet_name)
                parts.append(f"% {sheet_name}\n" + self._build_tabularx_table(headers, data, caption=sheet_name))
            content = "\n\n".join(parts)
        else:
            caption = self._current_sheet or self.file_path.stem
            content = self._build_tabularx_table(
                self._model.get_headers(), self._model.get_data(), caption=caption
            )

        self.convert_to_text.emit(content, f"{self.file_path.stem}.tex")

    # ── API pubblica ──────────────────────────────────────────────────────────

    def is_modified(self) -> bool:
        return self._model.is_modified()

    def save(self) -> bool:
        self._save()
        return not self._model.is_modified()
